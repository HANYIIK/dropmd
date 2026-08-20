import io
import re
import sys
from collections import Counter
from dataclasses import dataclass, field
from typing import Any, BinaryIO

from ._html_converter import HtmlConverter
from ._xlsx_semantic import (
    SemanticCell,
    SemanticSheet,
    SemanticWorkbook,
    build_semantic_workbook,
    has_value,
    plain_text,
)
from .._base_converter import DocumentConverter, DocumentConverterResult
from .._exceptions import MissingDependencyException, MISSING_DEPENDENCY_MESSAGE
from .._stream_info import StreamInfo

_xlsx_dependency_exc_info = None
try:
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
except ImportError:
    _xlsx_dependency_exc_info = sys.exc_info()

_xls_dependency_exc_info = None
try:
    import pandas as pd
    import xlrd  # noqa: F401
except ImportError:
    _xls_dependency_exc_info = sys.exc_info()

ACCEPTED_XLSX_MIME_TYPE_PREFIXES = [
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
]
ACCEPTED_XLSX_FILE_EXTENSIONS = [".xlsx"]

ACCEPTED_XLS_MIME_TYPE_PREFIXES = [
    "application/vnd.ms-excel",
    "application/excel",
]
ACCEPTED_XLS_FILE_EXTENSIONS = [".xls"]

_HIERARCHY_HEADER_PATTERN = re.compile(r"维度|层级|分类|模块|目录")
_IDENTIFIER_HEADER_PATTERN = re.compile(r"编号|序号|编码|(^|\b)id($|\b)", re.IGNORECASE)
_TOTAL_LABEL_PATTERN = re.compile(r"^(合计|小计|总计|总额|累计)$")
_FORMULA_ERROR_PATTERN = re.compile(r"^#(?:REF!|DIV/0!|VALUE!|NAME\?|N/A|NUM!|NULL!)$")
_LIST_ITEM_PATTERN = re.compile(
    r"^(?P<indent>[ \t]*)(?:(?P<number>\d+)(?P<number_punctuation>[.)、）])(?!\d)"
    r"|(?P<letter>[A-Za-z])(?P<letter_punctuation>[.)、）])"
    r"|(?P<bullet>[-*•●▪◦·]))[ \t]*(?P<text>\S.*)$"
)
_COLOR_EMOJI_PALETTE = (
    ("🟥", (255, 0, 0)),
    ("🟧", (255, 128, 0)),
    ("🟨", (255, 255, 0)),
    ("🟩", (0, 176, 80)),
    ("🟦", (0, 112, 192)),
    ("🟪", (112, 48, 160)),
    ("🟫", (150, 75, 0)),
    ("⬛", (0, 0, 0)),
    ("⬜", (255, 255, 255)),
)


@dataclass(frozen=True, slots=True)
class _StructuredLine:
    kind: str
    marker: str
    text: str
    level: int


@dataclass(slots=True)
class _SheetRenderResult:
    markdown: str
    source_cells: int
    represented_cells: int
    source_formulas: int
    represented_formulas: int
    source_cached_formulas: int
    represented_cached_formulas: int
    merge_categories: Counter[str] = field(default_factory=Counter)
    hidden_preserved: bool = False
    synthetic_rows: int = 0
    identifier_mutations: int = 0
    duplicate_identifier_groups: int = 0
    folded_rows: tuple[str, ...] = ()
    structured_cells: int = 0
    source_colored_cells: int = 0
    represented_colored_cells: int = 0


def _nearest_color_emoji(color: str) -> str:
    red, green, blue = (int(color[index : index + 2], 16) for index in (1, 3, 5))

    def distance(candidate: tuple[int, int, int]) -> float:
        candidate_red, candidate_green, candidate_blue = candidate
        red_mean = (red + candidate_red) / 2
        red_delta = red - candidate_red
        green_delta = green - candidate_green
        blue_delta = blue - candidate_blue
        return (
            (2 + red_mean / 256) * red_delta * red_delta
            + 4 * green_delta * green_delta
            + (2 + (255 - red_mean) / 256) * blue_delta * blue_delta
        )

    return min(_COLOR_EMOJI_PALETTE, key=lambda item: distance(item[1]))[0]


class _ColorLegend:
    def __init__(self, colors: tuple[str, ...]):
        groups: dict[str, list[str]] = {}
        for color in colors:
            groups.setdefault(_nearest_color_emoji(color), []).append(color)
        self.colors = colors
        self.markers: dict[str, str] = {}
        for emoji, group in groups.items():
            for index, color in enumerate(group, start=1):
                self.markers[color] = emoji if len(group) == 1 else f"{emoji}[C{index}]"

    def marker(self, color: str | None) -> str:
        return self.markers.get(color or "", "")

    def render(self) -> str:
        if not self.colors:
            return ""
        lines = [f"- {self.markers[color]} = `{color}`" for color in self.colors]
        return (
            "## Excel 单元格颜色图例\n\n"
            + "\n".join(lines)
            + "\n\n> Emoji 是近似视觉标记，精确填充色以十六进制值为准；不推断颜色的业务含义。"
        )


def _escape_markdown_text(value: Any) -> str:
    return plain_text(value).replace("\\", "\\\\").replace("|", "\\|")


def _inline_plain_text(value: Any) -> str:
    lines = [line.strip() for line in plain_text(value).split("\n") if line.strip()]
    if not lines:
        return ""
    text = lines[0]
    for line in lines[1:]:
        punctuation = ("。", "；", ";", "！", "？", "!", "?", "：", ":", "，", ",", ".")
        if line.startswith(("（", "(", "【", "[")) or text.endswith(punctuation):
            separator = " "
        else:
            separator = "； "
        text += separator + line
    return text


def _inline_markdown_text(value: Any) -> str:
    return _escape_markdown_text(_inline_plain_text(value))


def _structured_markdown_text(value: Any) -> str | None:
    source_lines = [line.rstrip() for line in plain_text(value).split("\n") if line.strip()]
    if len(source_lines) < 2:
        return None

    parsed: list[tuple[str, str, str, int]] = []
    marker_count = 0
    numbered_count = 0
    bullet_count = 0
    for source_line in source_lines:
        match = _LIST_ITEM_PATTERN.match(source_line)
        if match is None:
            parsed.append(("paragraph", "", source_line.strip(), 0))
            continue
        marker_count += 1
        indent = len(match.group("indent").expandtabs(4))
        if match.group("number"):
            kind = "number"
            marker = match.group("number")
            numbered_count += 1
        elif match.group("letter"):
            kind = "letter"
            marker = match.group("letter")
            numbered_count += 1
        else:
            kind = "bullet"
            marker = "-"
            bullet_count += 1
        parsed.append((kind, marker, match.group("text").strip(), indent))

    if marker_count < 2 or not (numbered_count >= 2 or bullet_count >= 2 or (numbered_count and bullet_count)):
        return None

    structured_lines: list[_StructuredLine] = []
    has_top_level_number = False
    previous_level = 0
    for kind, marker, text, indent in parsed:
        explicit_level = min(indent // 2, 3)
        if kind in {"number", "letter"}:
            level = explicit_level
            if kind == "letter" and has_top_level_number and explicit_level == 0:
                level = 1
            if kind == "number" and level == 0:
                has_top_level_number = True
        elif kind == "bullet":
            level = explicit_level or (1 if has_top_level_number else 0)
        else:
            level = previous_level + 1 if structured_lines and structured_lines[-1].kind != "paragraph" else 0
        structured_lines.append(_StructuredLine(kind=kind, marker=marker, text=text, level=level))
        previous_level = level

    rendered: list[str] = []
    for line in structured_lines:
        indent = "   " * line.level
        text = _escape_markdown_text(line.text)
        if line.kind in {"number", "letter"}:
            rendered.append(f"{indent}{line.marker}. {text}")
        elif line.kind == "bullet":
            rendered.append(f"{indent}- {text}")
        else:
            rendered.append(f"{indent}{text}")
    return "\n".join(rendered)


class _SheetRenderer:
    def __init__(self, sheet: SemanticSheet, color_legend: _ColorLegend | None = None):
        self.sheet = sheet
        self.color_legend = color_legend
        self.bounds = sheet.bounds
        self.merges = list(sheet.merges)
        self.represented_values: set[str] = set()
        self.represented_hyperlinks: set[str] = set()
        self.represented_comments: set[str] = set()
        self.represented_formulas: set[str] = set()
        self.represented_cached_formulas: set[str] = set()
        self.represented_colored_cells: set[str] = set()
        self.structured_details: dict[str, tuple[str, str]] = {}
        self.identifier_mutations = 0
        self.synthetic_rows = 0
        self.merge_by_cell: dict[tuple[int, int], Any] = {}
        for merged_range in self.merges:
            for row in range(merged_range.min_row, merged_range.max_row + 1):
                for column in range(merged_range.min_column, merged_range.max_column + 1):
                    self.merge_by_cell[(row, column)] = merged_range

    def _color_prefix(self, cell: SemanticCell) -> str:
        if self.color_legend is None or not cell.display_value or not cell.fill_rgb:
            return ""
        marker = self.color_legend.marker(cell.fill_rgb)
        if not marker:
            return ""
        self.represented_colored_cells.add(cell.coordinate)
        return f"{marker} "

    def _resolved_cell(self, row: int, column: int, *, expand_merges: bool = True) -> SemanticCell | None:
        merged_range = self.merge_by_cell.get((row, column))
        if merged_range is not None:
            anchor = (merged_range.min_row, merged_range.min_column)
            if not expand_merges and (row, column) != anchor:
                return None
            if merged_range.width > 1 and column != merged_range.min_column:
                return None
            row, column = anchor
        return self.sheet.cell(row, column)

    def _cell_value(
        self,
        row: int,
        column: int,
        *,
        expand_merges: bool = True,
        preserve_newlines: bool = False,
    ) -> str:
        cell = self._resolved_cell(row, column, expand_merges=expand_merges)
        if cell is None:
            return ""
        if cell.display_value:
            self.represented_values.add(cell.coordinate)
        markdown_text = (
            _escape_markdown_text(cell.display_value)
            if preserve_newlines
            else _inline_markdown_text(cell.display_value)
        )
        if cell.hyperlink:
            safe_target = cell.hyperlink.replace(" ", "%20").replace(">", "%3E")
            self.represented_hyperlinks.add(cell.coordinate)
            markdown_text = f"[{markdown_text or '链接'}](<{safe_target}>)"
        return self._color_prefix(cell) + markdown_text

    def _row_context(self, row: int, target_column: int) -> str:
        assert self.bounds is not None
        _, min_col, _, _ = self.bounds
        context: list[str] = []
        for column in range(min_col, target_column):
            cell = self._resolved_cell(row, column)
            if cell is None or not cell.display_value:
                continue
            value = _inline_markdown_text(cell.display_value)
            if value and value not in context:
                context.append(value)
        return " / ".join(context[-5:])

    def _table_cell_value(self, row: int, column: int) -> str:
        cell = self._resolved_cell(row, column)
        if cell is None or not cell.display_value:
            return self._cell_value(row, column)
        structured = _structured_markdown_text(cell.display_value)
        if structured is None:
            return self._cell_value(row, column)

        self.represented_values.add(cell.coordinate)
        if cell.hyperlink:
            safe_target = cell.hyperlink.replace(" ", "%20").replace(">", "%3E")
            structured += f"\n\n- 来源链接：<{safe_target}>"
            self.represented_hyperlinks.add(cell.coordinate)
        self.structured_details.setdefault(cell.coordinate, (self._row_context(row, column), structured))
        return self._color_prefix(cell) + f"详见「{cell.coordinate} 结构化详情」"

    def _full_width_bands(self) -> tuple[dict[int, Any], set[int]]:
        if self.bounds is None:
            return {}, set()
        _, min_col, _, max_col = self.bounds
        bands: dict[int, Any] = {}
        covered_rows: set[int] = set()
        for merged_range in self.merges:
            if merged_range.min_column <= min_col and merged_range.max_column >= max_col:
                value = self._cell_value(merged_range.min_row, merged_range.min_column)
                if value:
                    bands[merged_range.min_row] = merged_range
                    covered_rows.update(range(merged_range.min_row, merged_range.max_row + 1))
        return bands, covered_rows

    def _header_row(self, band_rows: set[int]) -> int | None:
        if self.bounds is None:
            return None
        min_row, min_col, max_row, max_col = self.bounds
        fallback = None
        for row in range(min_row, max_row + 1):
            if row in band_rows:
                continue
            values = [self._cell_value(row, column, expand_merges=False) for column in range(min_col, max_col + 1)]
            count = sum(bool(value) for value in values)
            if count and fallback is None:
                fallback = row
            if count >= 2:
                return row
        return fallback

    def _headers(self, header_row: int) -> list[str]:
        assert self.bounds is not None
        _, min_col, _, max_col = self.bounds
        headers: list[str] = []
        for column in range(min_col, max_col + 1):
            merged_range = self.merge_by_cell.get((header_row, column))
            value = self._cell_value(header_row, column)
            if merged_range is not None and merged_range.width > 1:
                base = self._cell_value(merged_range.min_row, merged_range.min_column)
                position = column - merged_range.min_column + 1
                suffix = f"{position}级" if _HIERARCHY_HEADER_PATTERN.search(base) else str(position)
                value = f"{base}（{suffix}）"
            headers.append(value or f"列 {get_column_letter(column)}")

        counts: dict[str, int] = {}
        unique_headers: list[str] = []
        for header in headers:
            counts[header] = counts.get(header, 0) + 1
            occurrence = counts[header]
            unique_headers.append(header if occurrence == 1 else f"{header}（{occurrence}）")
        return unique_headers

    @staticmethod
    def _render_band(value: str, height: int) -> str:
        plain = value
        if height == 1 and len(plain) <= 80 and "\n" not in plain:
            return f"### {plain}"
        lines = plain.splitlines() or [plain]
        return "\n".join(f"> {line}" if line else ">" for line in lines)

    @staticmethod
    def _render_table(headers: list[str], rows: list[list[str]]) -> str:
        output = [
            "| " + " | ".join(headers) + " |",
            "| " + " | ".join("---" for _ in headers) + " |",
        ]
        output.extend("| " + " | ".join(row) + " |" for row in rows)
        return "\n".join(output)

    def _is_total_row(self, row: int) -> bool:
        assert self.bounds is not None
        _, min_col, _, max_col = self.bounds
        for column in range(min_col, max_col + 1):
            cell = self._resolved_cell(row, column)
            if cell is not None and _TOTAL_LABEL_PATTERN.fullmatch(cell.display_value.strip()):
                return True
        return False

    def _has_direct_source_content(self, row: int) -> bool:
        assert self.bounds is not None
        _, min_col, _, max_col = self.bounds
        return any(self.sheet.cell(row, column) is not None for column in range(min_col, max_col + 1))

    def _is_merge_continuation(self, row: int) -> bool:
        assert self.bounds is not None
        _, min_col, _, max_col = self.bounds
        return any(
            merged_range.min_row < row <= merged_range.max_row
            and merged_range.min_column <= max_col
            and merged_range.max_column >= min_col
            for merged_range in self.merges
        )

    def _source_warnings(
        self, header_row: int, headers: list[str], skipped_rows: set[int]
    ) -> tuple[list[str], int]:
        assert self.bounds is not None
        _, min_col, max_row, _ = self.bounds
        warnings: list[str] = []
        duplicate_group_count = 0
        for offset, header in enumerate(headers):
            if not _IDENTIFIER_HEADER_PATTERN.search(header):
                continue
            column = min_col + offset
            groups: list[dict[str, list[str]]] = []
            occurrences: dict[str, list[str]] = {}
            for row in range(header_row + 1, max_row + 1):
                if row in skipped_rows:
                    if occurrences:
                        groups.append(occurrences)
                        occurrences = {}
                    continue
                cell = self.sheet.cell(row, column)
                if cell is None or not has_value(cell.source_value):
                    continue
                value = _inline_markdown_text(cell.display_value)
                if value:
                    occurrences.setdefault(value, []).append(cell.coordinate)
            if occurrences:
                groups.append(occurrences)
            duplicates = [
                (value, coordinates)
                for group in groups
                for value, coordinates in group.items()
                if len(coordinates) > 1
            ]
            if duplicates:
                duplicate_group_count += len(duplicates)
                details = "；".join(
                    f"`{value.replace('`', '´')}`（{'、'.join(coordinates)}）"
                    for value, coordinates in duplicates[:20]
                )
                if len(duplicates) > 20:
                    details += f"；另有 {len(duplicates) - 20} 项"
                warnings.append(
                    f"> **源数据提示：** “{header}”列存在重复值：{details}。DropMD 已按原文件保留。"
                )
        return warnings, duplicate_group_count

    def _comments(self) -> list[str]:
        comments = []
        for cell in sorted(self.sheet.cells.values(), key=lambda item: (item.row, item.column)):
            if cell.comment:
                author = f"（{_inline_markdown_text(cell.comment_author)}）" if cell.comment_author else ""
                comments.append(f"- `{cell.coordinate}`{author}：{_inline_markdown_text(cell.comment)}")
                self.represented_comments.add(cell.coordinate)
        return comments

    def _formulas(self) -> tuple[list[str], list[str]]:
        formulas: list[str] = []
        errors: list[str] = []
        for cell in sorted(self.sheet.formula_cells, key=lambda item: (item.row, item.column)):
            formula = cell.formula.replace("`", "´") if cell.formula else ""
            if cell.has_cached_value:
                result = _inline_markdown_text(cell.display_value).replace("`", "´")
                formulas.append(f"- `{cell.coordinate}`：`{formula}` → `{result}`")
                self.represented_cached_formulas.add(cell.coordinate)
                if _FORMULA_ERROR_PATTERN.fullmatch(cell.display_value.strip()):
                    errors.append(f"`{cell.coordinate}`（`{cell.display_value}`）")
            else:
                formulas.append(f"- `{cell.coordinate}`：`{formula}`（原文件未保存计算结果）")
            self.represented_formulas.add(cell.coordinate)
        return formulas, errors

    def _merge_categories(self, header_row: int | None) -> Counter[str]:
        categories: Counter[str] = Counter()
        if self.bounds is None:
            return categories
        _, min_col, _, max_col = self.bounds
        for merged_range in self.merges:
            if merged_range.min_column <= min_col and merged_range.max_column >= max_col:
                categories["章节或说明"] += 1
            elif header_row is not None and merged_range.min_row <= header_row <= merged_range.max_row:
                categories["表头"] += 1
            elif merged_range.height > 1 and merged_range.width == 1:
                categories["纵向层级或分组"] += 1
            else:
                categories["其他布局"] += 1
        return categories

    def _check_identifier_value(self, row: int, column: int) -> None:
        cell = self.sheet.cell(row, column)
        if cell is None or not cell.display_value:
            return
        if plain_text(cell.display_value) != _inline_plain_text(cell.display_value):
            self.identifier_mutations += 1

    def _represented_cell_count(self) -> int:
        represented = 0
        for cell in self.sheet.cells.values():
            value_ok = not cell.display_value or cell.coordinate in self.represented_values
            hyperlink_ok = not cell.hyperlink or cell.coordinate in self.represented_hyperlinks
            comment_ok = not cell.comment or cell.coordinate in self.represented_comments
            formula_ok = not cell.formula or cell.coordinate in self.represented_formulas
            if value_ok and hyperlink_ok and comment_ok and formula_ok:
                represented += 1
        return represented

    def _result(
        self,
        parts: list[str],
        header_row: int | None,
        folded_rows: list[int] | None = None,
        duplicate_identifier_groups: int = 0,
    ) -> _SheetRenderResult:
        folded_rows = folded_rows or []
        cached_formulas = {cell.coordinate for cell in self.sheet.formula_cells if cell.has_cached_value}
        colored_cells = {
            cell.coordinate for cell in self.sheet.cells.values() if cell.display_value and cell.fill_rgb
        }
        return _SheetRenderResult(
            markdown="\n\n".join(part for part in parts if part),
            source_cells=len(self.sheet.cells),
            represented_cells=self._represented_cell_count(),
            source_formulas=len(self.sheet.formula_cells),
            represented_formulas=len(self.represented_formulas),
            source_cached_formulas=len(cached_formulas),
            represented_cached_formulas=len(cached_formulas & self.represented_cached_formulas),
            merge_categories=self._merge_categories(header_row),
            hidden_preserved=self.sheet.state != "visible",
            synthetic_rows=self.synthetic_rows,
            identifier_mutations=self.identifier_mutations,
            duplicate_identifier_groups=duplicate_identifier_groups,
            folded_rows=tuple(f"{self.sheet.name}!{row}" for row in folded_rows),
            structured_cells=len(self.structured_details),
            source_colored_cells=len(colored_cells),
            represented_colored_cells=len(colored_cells & self.represented_colored_cells),
        )

    def _source_note(self) -> str | None:
        details: list[str] = []
        if self.sheet.state != "visible":
            state = "隐藏" if self.sheet.state == "hidden" else "深度隐藏"
            details.append(f"原工作表为{state}状态")
        if self.sheet.hidden_rows:
            details.append(f"隐藏行 {len(self.sheet.hidden_rows)} 个")
        if self.sheet.hidden_columns:
            details.append(f"隐藏列 {len(self.sheet.hidden_columns)} 个")
        if not details:
            return None
        return "> **源表提示：** " + "；".join(details) + "。"

    def render(self) -> _SheetRenderResult:
        title = _inline_markdown_text(self.sheet.name)
        parts = [f"## {title}"]
        source_note = self._source_note()
        if source_note:
            parts.append(source_note)
        if self.bounds is None:
            return self._result(parts, None)

        min_row, min_col, max_row, max_col = self.bounds
        bands, band_rows = self._full_width_bands()
        header_row = self._header_row(band_rows)
        if header_row is None:
            comments = self._comments()
            if comments:
                parts.append("### 批注\n\n" + "\n".join(comments))
            formulas, _ = self._formulas()
            if formulas:
                parts.append("### 公式\n\n" + "\n".join(formulas))
            return self._result(parts, None)

        for row in range(min_row, header_row):
            if row in bands:
                merged_range = bands[row]
                value = self._cell_value(row, merged_range.min_column, preserve_newlines=True)
                if value and value.replace("\n", " ") != title:
                    parts.append(self._render_band(value, merged_range.height))
            elif row not in band_rows:
                if not self._has_direct_source_content(row):
                    continue
                values = [self._cell_value(row, column) for column in range(min_col, max_col + 1)]
                values = [value for value in values if value]
                if values:
                    parts.append("  ".join(values))

        headers = self._headers(header_row)
        identifier_columns = {
            min_col + offset for offset, header in enumerate(headers) if _IDENTIFIER_HEADER_PATTERN.search(header)
        }
        table_rows: list[list[str]] = []
        folded_rows: list[int] = []

        def flush_table() -> None:
            if table_rows:
                parts.append(self._render_table(headers, table_rows))
                table_rows.clear()

        for row in range(header_row + 1, max_row + 1):
            if row in bands:
                flush_table()
                merged_range = bands[row]
                value = self._cell_value(row, merged_range.min_column, preserve_newlines=True)
                parts.append(self._render_band(value, merged_range.height))
                continue
            if row in band_rows:
                continue
            if not self._has_direct_source_content(row):
                if self._is_merge_continuation(row):
                    folded_rows.append(row)
                else:
                    flush_table()
                continue
            values = [self._table_cell_value(row, column) for column in range(min_col, max_col + 1)]
            if any(values):
                for column in identifier_columns:
                    self._check_identifier_value(row, column)
                if self._is_total_row(row):
                    values = [f"**{value}**" if value else "" for value in values]
                table_rows.append(values)
            else:
                flush_table()
        flush_table()

        if self.structured_details:
            details = ["### 结构化单元格详情"]
            for coordinate, (context, structured) in self.structured_details.items():
                heading = f"#### {coordinate}"
                if context:
                    heading += f" · {context}"
                details.append(f"{heading}\n\n{structured}")
            parts.append("\n\n".join(details))

        if folded_rows:
            row_list = "、".join(str(row) for row in folded_rows[:20])
            if len(folded_rows) > 20:
                row_list += f"，另有 {len(folded_rows) - 20} 行"
            parts.append(
                f"> **转换提示：** 已折叠 {len(folded_rows)} 个仅用于合并单元格排版的延续行"
                f"（源表第 {row_list} 行），未生成重复记录。"
            )
        warnings, duplicate_identifier_groups = self._source_warnings(header_row, headers, band_rows)
        parts.extend(warnings)
        comments = self._comments()
        if comments:
            parts.append("### 批注\n\n" + "\n".join(comments))
        formulas, formula_errors = self._formulas()
        if formulas:
            parts.append("### 公式\n\n" + "\n".join(formulas))
        if formula_errors:
            parts.append(
                "> **公式提示：** 以下单元格包含错误结果：" + "、".join(formula_errors) + "。"
            )
        return self._result(parts, header_row, folded_rows, duplicate_identifier_groups)


class _WorkbookRenderer:
    def __init__(self, workbook: SemanticWorkbook, preserve_excel_colors: bool = False):
        self.workbook = workbook
        self.preserve_excel_colors = preserve_excel_colors
        self.color_legend = _ColorLegend(workbook.fill_colors) if preserve_excel_colors else None

    @staticmethod
    def _state_label(state: str) -> str:
        if state == "visible":
            return "可见"
        if state == "hidden":
            return "隐藏"
        return "深度隐藏"

    def render(self) -> str:
        title = _inline_markdown_text(self.workbook.title or "Excel 工作簿")
        visible_count = sum(sheet.state == "visible" for sheet in self.workbook.sheets)
        hidden_count = len(self.workbook.sheets) - visible_count
        rendered_sheets = [
            _SheetRenderer(sheet, self.color_legend).render() for sheet in self.workbook.sheets
        ]
        source_cells = sum(result.source_cells for result in rendered_sheets)
        represented_cells = sum(result.represented_cells for result in rendered_sheets)
        source_formulas = sum(result.source_formulas for result in rendered_sheets)
        represented_formulas = sum(result.represented_formulas for result in rendered_sheets)
        source_cached_formulas = sum(result.source_cached_formulas for result in rendered_sheets)
        represented_cached_formulas = sum(result.represented_cached_formulas for result in rendered_sheets)
        merge_categories: Counter[str] = Counter()
        for result in rendered_sheets:
            merge_categories.update(result.merge_categories)
        classified_merges = sum(merge_categories.values())
        preserved_hidden = sum(result.hidden_preserved for result in rendered_sheets)
        synthetic_rows = sum(result.synthetic_rows for result in rendered_sheets)
        identifier_mutations = sum(result.identifier_mutations for result in rendered_sheets)
        duplicate_identifier_groups = sum(result.duplicate_identifier_groups for result in rendered_sheets)
        folded_rows = [coordinate for result in rendered_sheets for coordinate in result.folded_rows]
        structured_cells = sum(result.structured_cells for result in rendered_sheets)
        source_colored_cells = sum(result.source_colored_cells for result in rendered_sheets)
        represented_colored_cells = sum(result.represented_colored_cells for result in rendered_sheets)
        overview = (
            f"> **工作簿概览：** {len(self.workbook.sheets)} 个工作表"
            f"（{visible_count} 个可见，{hidden_count} 个隐藏）；"
            f"{self.workbook.merge_count} 个合并区域；{self.workbook.formula_count} 个公式单元格。"
        )
        index = []
        for sheet in self.workbook.sheets:
            details = [f"区域 `{sheet.dimension or '空白'}`", self._state_label(sheet.state)]
            if sheet.merges:
                details.append(f"{len(sheet.merges)} 个合并区域")
            if sheet.formula_cells:
                details.append(f"{len(sheet.formula_cells)} 个公式")
            if sheet.table_refs:
                details.append(f"{len(sheet.table_refs)} 个结构化表")
            index.append(f"- `{_inline_markdown_text(sheet.name)}`：" + "；".join(details))

        category_text = "；".join(
            f"{name} {merge_categories[name]}"
            for name in ("纵向层级或分组", "章节或说明", "表头", "其他布局")
            if merge_categories[name]
        )
        folded_text = "、".join(f"`{coordinate}`" for coordinate in folded_rows[:20])
        if len(folded_rows) > 20:
            folded_text += f"，另有 {len(folded_rows) - 20} 行"
        integrity_lines = [
            f"- 工作表：{len(rendered_sheets)}/{len(self.workbook.sheets)} 已保留"
            f"（隐藏工作表 {preserved_hidden}/{hidden_count}）",
            f"- 可追踪语义单元格：{represented_cells}/{source_cells} 已表示",
            f"- 公式：{represented_formulas}/{source_formulas} 已保留"
            f"（缓存结果 {represented_cached_formulas}/{source_cached_formulas}）",
            f"- 合并区域：{classified_merges}/{self.workbook.merge_count} 已分类"
            + (f"（{category_text}）" if category_text else ""),
            f"- 结构化列表单元格：{structured_cells} 个已保留为 Markdown 层级",
            f"- 合成记录：{synthetic_rows}",
            f"- 编号显示值变更：{identifier_mutations}",
            f"- 源数据异常：{duplicate_identifier_groups} 组重复编号（按原值保留）",
            f"- 折叠排版延续行：{len(folded_rows)}" + (f"（{folded_text}）" if folded_text else ""),
        ]
        if self.preserve_excel_colors:
            integrity_lines.insert(
                4,
                f"- 单元格填充色：{represented_colored_cells}/{source_colored_cells} 个非空着色单元格已标记"
                f"（{len(self.workbook.fill_colors)} 种精确颜色）",
            )
        integrity_note = (
            "> 本摘要核对语义内容与来源关系，不代表像素级还原字体、边框或版式；"
            "颜色 Emoji 为近似显示，精确值见图例。"
            if self.preserve_excel_colors
            else "> 本摘要核对语义内容与来源关系，不代表像素级还原字体、颜色或版式。"
        )
        integrity = (
            "## 语义转换完整性\n\n"
            + "\n".join(integrity_lines)
            + "\n\n"
            + integrity_note
        )
        parts = [f"# {title}", overview, integrity, "## 工作表索引\n\n" + "\n".join(index)]
        if self.color_legend is not None:
            parts.insert(3, self.color_legend.render())
        parts.extend(result.markdown for result in rendered_sheets)
        if self.workbook.defined_names:
            names = "、".join(f"`{_inline_markdown_text(name)}`" for name, _ in self.workbook.defined_names)
            parts.append(f"> **定义名称：** {names}。")
        return "\n\n".join(part for part in parts if part)


class XlsxConverter(DocumentConverter):
    """Converts XLSX files while preserving merged-cell hierarchy and text identifiers."""

    def accepts(
        self,
        file_stream: BinaryIO,
        stream_info: StreamInfo,
        **kwargs: Any,
    ) -> bool:
        mimetype = (stream_info.mimetype or "").lower()
        extension = (stream_info.extension or "").lower()
        return extension in ACCEPTED_XLSX_FILE_EXTENSIONS or any(
            mimetype.startswith(prefix) for prefix in ACCEPTED_XLSX_MIME_TYPE_PREFIXES
        )

    def convert(
        self,
        file_stream: BinaryIO,
        stream_info: StreamInfo,
        **kwargs: Any,
    ) -> DocumentConverterResult:
        if _xlsx_dependency_exc_info is not None:
            raise MissingDependencyException(
                MISSING_DEPENDENCY_MESSAGE.format(
                    converter=type(self).__name__,
                    extension=".xlsx",
                    feature="xlsx",
                )
            ) from _xlsx_dependency_exc_info[1].with_traceback(_xlsx_dependency_exc_info[2])  # type: ignore[union-attr]

        file_stream.seek(0)
        workbook_bytes = file_stream.read()
        workbook = load_workbook(io.BytesIO(workbook_bytes), data_only=False, keep_links=False)
        cached_workbook = load_workbook(io.BytesIO(workbook_bytes), data_only=True, keep_links=False)
        semantic_workbook = build_semantic_workbook(workbook, cached_workbook, stream_info.filename)
        preserve_excel_colors = bool(kwargs.get("preserve_excel_colors", False))
        return DocumentConverterResult(
            markdown=_WorkbookRenderer(semantic_workbook, preserve_excel_colors).render().strip(),
            title=semantic_workbook.title,
        )


class XlsConverter(DocumentConverter):
    """Converts legacy XLS files to Markdown tables."""

    def __init__(self):
        super().__init__()
        self._html_converter = HtmlConverter()

    def accepts(
        self,
        file_stream: BinaryIO,
        stream_info: StreamInfo,
        **kwargs: Any,
    ) -> bool:
        mimetype = (stream_info.mimetype or "").lower()
        extension = (stream_info.extension or "").lower()
        return extension in ACCEPTED_XLS_FILE_EXTENSIONS or any(
            mimetype.startswith(prefix) for prefix in ACCEPTED_XLS_MIME_TYPE_PREFIXES
        )

    def convert(
        self,
        file_stream: BinaryIO,
        stream_info: StreamInfo,
        **kwargs: Any,
    ) -> DocumentConverterResult:
        if _xls_dependency_exc_info is not None:
            raise MissingDependencyException(
                MISSING_DEPENDENCY_MESSAGE.format(
                    converter=type(self).__name__,
                    extension=".xls",
                    feature="xls",
                )
            ) from _xls_dependency_exc_info[1].with_traceback(_xls_dependency_exc_info[2])  # type: ignore[union-attr]

        sheets = pd.read_excel(file_stream, sheet_name=None, engine="xlrd")
        md_content = ""
        for sheet_name, sheet in sheets.items():
            md_content += f"## {sheet_name}\n"
            html_content = sheet.to_html(index=False)
            md_content += self._html_converter.convert_string(html_content, **kwargs).markdown.strip() + "\n\n"
        return DocumentConverterResult(markdown=md_content.strip())
