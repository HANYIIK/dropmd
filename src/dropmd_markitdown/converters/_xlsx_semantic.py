from __future__ import annotations

import colorsys
import re
import xml.etree.ElementTree as ElementTree
from dataclasses import dataclass
from datetime import date, datetime, time
from decimal import Decimal
from pathlib import Path
from typing import Any

def has_value(value: Any) -> bool:
    return value is not None and (not isinstance(value, str) or bool(value.strip()))


def plain_text(value: Any) -> str:
    return str(value).replace("\r\n", "\n").replace("\r", "\n").strip()


def _column_letter(column: int) -> str:
    result = ""
    while column:
        column, remainder = divmod(column - 1, 26)
        result = chr(65 + remainder) + result
    return result


def _format_number(value: int | float | Decimal, number_format: str) -> str:
    if isinstance(value, int):
        raw = str(value)
    else:
        raw = format(Decimal(str(value)), "f").rstrip("0").rstrip(".")
        if raw in {"", "-0"}:
            raw = "0"

    section = (number_format or "General").split(";", 1)[0]
    if section.lower() == "general" or not re.search(r"[0#]", section):
        return raw

    is_percent = "%" in section
    numeric_value = Decimal(str(value)) * (100 if is_percent else 1)
    placeholder = re.sub(r'"[^"]*"|\[[^\]]*\]|\\.', "", section)
    scientific = re.search(r"([0#]+)(?:\.([0#]+))?E[+-]0+", placeholder, re.IGNORECASE)
    if scientific:
        decimals = len(scientific.group(2) or "")
        rendered = f"{float(numeric_value):.{decimals}E}"
    else:
        decimal_match = re.search(r"\.([0#]+)", placeholder)
        decimal_pattern = decimal_match.group(1) if decimal_match else ""
        minimum_decimals = decimal_pattern.count("0")
        maximum_decimals = len(decimal_pattern)
        grouping = "," in placeholder.split(".", 1)[0]
        if maximum_decimals:
            rendered = f"{numeric_value:,.{maximum_decimals}f}" if grouping else f"{numeric_value:.{maximum_decimals}f}"
            if maximum_decimals > minimum_decimals:
                integer_part, fraction = rendered.split(".", 1)
                fraction = fraction.rstrip("0")
                if len(fraction) < minimum_decimals:
                    fraction += "0" * (minimum_decimals - len(fraction))
                rendered = integer_part + (f".{fraction}" if fraction else "")
        else:
            rendered = f"{numeric_value:,.0f}" if grouping else f"{numeric_value:.0f}"
    return rendered + ("%" if is_percent else "")


def _display_scalar(value: Any, number_format: str) -> str:
    if not has_value(value):
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, datetime):
        return value.isoformat(sep=" ", timespec="seconds")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, time):
        return value.isoformat(timespec="seconds")
    if isinstance(value, (int, float, Decimal)):
        return _format_number(value, number_format)
    return plain_text(value)


@dataclass(frozen=True, slots=True)
class SemanticCell:
    row: int
    column: int
    coordinate: str
    source_value: Any
    cached_value: Any
    display_value: str
    formula: str | None
    number_format: str
    hyperlink: str | None
    comment: str | None
    comment_author: str | None
    is_bold: bool
    fill_rgb: str | None

    @property
    def has_cached_value(self) -> bool:
        return self.formula is not None and has_value(self.cached_value)


@dataclass(frozen=True, slots=True)
class SemanticMerge:
    ref: str
    min_row: int
    min_column: int
    max_row: int
    max_column: int

    @property
    def width(self) -> int:
        return self.max_column - self.min_column + 1

    @property
    def height(self) -> int:
        return self.max_row - self.min_row + 1


@dataclass(frozen=True, slots=True)
class SemanticSheet:
    name: str
    state: str
    min_row: int | None
    min_column: int | None
    max_row: int | None
    max_column: int | None
    cells: dict[tuple[int, int], SemanticCell]
    merges: tuple[SemanticMerge, ...]
    hidden_rows: tuple[int, ...]
    hidden_columns: tuple[str, ...]
    table_refs: tuple[str, ...]
    freeze_panes: str | None
    auto_filter_ref: str | None

    @property
    def bounds(self) -> tuple[int, int, int, int] | None:
        if None in (self.min_row, self.min_column, self.max_row, self.max_column):
            return None
        return self.min_row, self.min_column, self.max_row, self.max_column  # type: ignore[return-value]

    @property
    def dimension(self) -> str | None:
        if self.bounds is None:
            return None
        min_row, min_column, max_row, max_column = self.bounds
        return f"{_column_letter(min_column)}{min_row}:{_column_letter(max_column)}{max_row}"

    @property
    def formula_cells(self) -> tuple[SemanticCell, ...]:
        return tuple(cell for cell in self.cells.values() if cell.formula is not None)

    def cell(self, row: int, column: int) -> SemanticCell | None:
        return self.cells.get((row, column))


@dataclass(frozen=True, slots=True)
class SemanticWorkbook:
    title: str | None
    sheets: tuple[SemanticSheet, ...]
    defined_names: tuple[tuple[str, str], ...]

    @property
    def merge_count(self) -> int:
        return sum(len(sheet.merges) for sheet in self.sheets)

    @property
    def formula_count(self) -> int:
        return sum(len(sheet.formula_cells) for sheet in self.sheets)

    @property
    def fill_colors(self) -> tuple[str, ...]:
        colors: list[str] = []
        seen: set[str] = set()
        for sheet in self.sheets:
            for cell in sorted(sheet.cells.values(), key=lambda item: (item.row, item.column)):
                if cell.display_value and cell.fill_rgb and cell.fill_rgb not in seen:
                    seen.add(cell.fill_rgb)
                    colors.append(cell.fill_rgb)
        return tuple(colors)


def _normalize_rgb(value: Any) -> str | None:
    text = str(value or "").strip().lstrip("#").upper()
    if len(text) == 8:
        text = text[-6:]
    if len(text) != 6 or not re.fullmatch(r"[0-9A-F]{6}", text):
        return None
    return f"#{text}"


def _theme_colors(workbook: Any) -> tuple[str, ...]:
    theme = getattr(workbook, "loaded_theme", None)
    if not theme:
        return ()
    try:
        root = ElementTree.fromstring(theme)
        scheme = root.find(".//{http://schemas.openxmlformats.org/drawingml/2006/main}clrScheme")
    except (ElementTree.ParseError, TypeError, ValueError):
        return ()
    if scheme is None:
        return ()
    colors: list[str] = []
    for entry in scheme:
        color_node = next(iter(entry), None)
        value = None if color_node is None else color_node.attrib.get("lastClr") or color_node.attrib.get("val")
        normalized = _normalize_rgb(value)
        if normalized:
            colors.append(normalized)
    return tuple(colors)


def _apply_tint(rgb: str, tint: float) -> str:
    if not tint:
        return rgb
    red, green, blue = (int(rgb[index : index + 2], 16) / 255 for index in (1, 3, 5))
    hue, lightness, saturation = colorsys.rgb_to_hls(red, green, blue)
    if tint < 0:
        lightness *= 1 + tint
    else:
        lightness = lightness * (1 - tint) + tint
    tinted = colorsys.hls_to_rgb(hue, max(0.0, min(1.0, lightness)), saturation)
    return "#" + "".join(f"{round(channel * 255):02X}" for channel in tinted)


def _resolve_fill_rgb(cell: Any, theme_colors: tuple[str, ...]) -> str | None:
    fill = getattr(cell, "fill", None)
    if fill is None or getattr(fill, "fill_type", None) != "solid":
        return None
    color = getattr(fill, "fgColor", None)
    if color is None:
        return None
    color_type = getattr(color, "type", None)
    resolved: str | None = None
    if color_type == "rgb":
        resolved = _normalize_rgb(getattr(color, "rgb", None))
    elif color_type == "theme":
        theme_index = getattr(color, "theme", None)
        if isinstance(theme_index, int) and 0 <= theme_index < len(theme_colors):
            resolved = theme_colors[theme_index]
    elif color_type == "indexed":
        indexed = getattr(color, "indexed", None)
        try:
            from openpyxl.styles.colors import COLOR_INDEX

            if isinstance(indexed, int) and 0 <= indexed < len(COLOR_INDEX):
                resolved = _normalize_rgb(COLOR_INDEX[indexed])
        except ImportError:
            resolved = None
    if resolved is None:
        return None
    return _apply_tint(resolved, float(getattr(color, "tint", 0.0) or 0.0))


def _sheet_cells(
    sheet: Any,
    cached_sheet: Any,
    theme_colors: tuple[str, ...],
) -> dict[tuple[int, int], SemanticCell]:
    cells: dict[tuple[int, int], SemanticCell] = {}
    for source_cell in sheet._cells.values():
        comment = getattr(source_cell, "comment", None)
        hyperlink = getattr(source_cell, "hyperlink", None)
        target = getattr(hyperlink, "target", None) or getattr(hyperlink, "location", None)
        formula = str(source_cell.value) if getattr(source_cell, "data_type", None) == "f" else None
        if not has_value(source_cell.value) and comment is None and target is None:
            continue

        cached_cell = cached_sheet.cell(source_cell.row, source_cell.column)
        cached_value = cached_cell.value if formula is not None else None
        displayed_value = cached_value if formula is not None and has_value(cached_value) else source_cell.value
        format_cell = cached_cell if formula is not None and has_value(cached_value) else source_cell
        cells[(source_cell.row, source_cell.column)] = SemanticCell(
            row=source_cell.row,
            column=source_cell.column,
            coordinate=source_cell.coordinate,
            source_value=source_cell.value,
            cached_value=cached_value,
            display_value=_display_scalar(displayed_value, getattr(format_cell, "number_format", "General")),
            formula=formula,
            number_format=getattr(source_cell, "number_format", "General"),
            hyperlink=str(target) if target else None,
            comment=plain_text(comment.text) if comment is not None and has_value(comment.text) else None,
            comment_author=plain_text(comment.author) if comment is not None and has_value(comment.author) else None,
            is_bold=bool(getattr(getattr(source_cell, "font", None), "bold", False)),
            fill_rgb=_resolve_fill_rgb(source_cell, theme_colors),
        )
    return cells


def _sheet_bounds(sheet: Any, cells: dict[tuple[int, int], SemanticCell]) -> tuple[int, int, int, int] | None:
    if not cells:
        return None
    rows = [row for row, _ in cells]
    columns = [column for _, column in cells]
    min_row, max_row = min(rows), max(rows)
    min_column, max_column = min(columns), max(columns)
    for merged_range in sheet.merged_cells.ranges:
        if (merged_range.min_row, merged_range.min_col) in cells:
            min_row = min(min_row, merged_range.min_row)
            min_column = min(min_column, merged_range.min_col)
            max_row = max(max_row, merged_range.max_row)
            max_column = max(max_column, merged_range.max_col)
    return min_row, min_column, max_row, max_column


def build_semantic_workbook(workbook: Any, cached_workbook: Any, filename: str | None = None) -> SemanticWorkbook:
    sheets: list[SemanticSheet] = []
    theme_colors = _theme_colors(workbook)
    for sheet in workbook.worksheets:
        cells = _sheet_cells(sheet, cached_workbook[sheet.title], theme_colors)
        bounds = _sheet_bounds(sheet, cells)
        merges = tuple(
            SemanticMerge(
                ref=str(merged_range),
                min_row=merged_range.min_row,
                min_column=merged_range.min_col,
                max_row=merged_range.max_row,
                max_column=merged_range.max_col,
            )
            for merged_range in sheet.merged_cells.ranges
        )
        freeze_panes = getattr(sheet, "freeze_panes", None)
        if freeze_panes is not None:
            freeze_panes = getattr(freeze_panes, "coordinate", str(freeze_panes))
        tables = tuple(table.ref for table in sheet.tables.values())
        auto_filter_ref = getattr(getattr(sheet, "auto_filter", None), "ref", None)
        sheets.append(
            SemanticSheet(
                name=sheet.title,
                state=sheet.sheet_state,
                min_row=bounds[0] if bounds else None,
                min_column=bounds[1] if bounds else None,
                max_row=bounds[2] if bounds else None,
                max_column=bounds[3] if bounds else None,
                cells=cells,
                merges=merges,
                hidden_rows=tuple(sorted(row for row, dimension in sheet.row_dimensions.items() if dimension.hidden)),
                hidden_columns=tuple(
                    sorted(column for column, dimension in sheet.column_dimensions.items() if dimension.hidden)
                ),
                table_refs=tables,
                freeze_panes=str(freeze_panes) if freeze_panes else None,
                auto_filter_ref=str(auto_filter_ref) if auto_filter_ref else None,
            )
        )

    defined_names = tuple(
        (name, str(defined_name.attr_text))
        for name, defined_name in workbook.defined_names.items()
        if getattr(defined_name, "attr_text", None)
    )
    title = Path(filename).stem if filename else None
    return SemanticWorkbook(title=title, sheets=tuple(sheets), defined_names=defined_names)
