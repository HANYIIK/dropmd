from pathlib import Path
from copy import copy
import re
from types import SimpleNamespace
from zipfile import ZIP_DEFLATED, ZipFile

import pytest
from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill

from dropmd.converter import OutputExistsError, UnsupportedFileError, convert_file, output_path_for


class FakeConverter:
    def convert(self, source: Path):
        return SimpleNamespace(markdown=f"# {source.stem}\n\n转换成功")


def set_formula_cache(source: Path, coordinate: str, value: str) -> None:
    patched = source.with_name(f"{source.stem}-cached.xlsx")
    with ZipFile(source) as original, ZipFile(patched, "w", ZIP_DEFLATED) as updated:
        for item in original.infolist():
            data = original.read(item.filename)
            if item.filename == "xl/worksheets/sheet1.xml":
                xml = data.decode("utf-8")
                pattern = rf'(<c r="{re.escape(coordinate)}"[^>]*><f>.*?</f><v>).*?(</v>)'
                xml, count = re.subn(pattern, rf"\g<1>{value}\g<2>", xml, count=1)
                assert count == 1
                data = xml.encode("utf-8")
            updated.writestr(item, data)
    patched.replace(source)


def test_output_path_keeps_the_original_stem(tmp_path: Path):
    source = tmp_path / "AI 打标字段 (2).docx"
    assert output_path_for(source) == tmp_path / "AI 打标字段 (2).md"


def test_convert_writes_utf8_markdown_beside_source(tmp_path: Path):
    source = tmp_path / "示例.docx"
    source.write_bytes(b"test")

    destination = convert_file(source, converter_factory=FakeConverter)

    assert destination == tmp_path / "示例.md"
    assert destination.read_text(encoding="utf-8") == "# 示例\n\n转换成功\n"


def test_excel_color_option_is_only_forwarded_for_xlsx(tmp_path: Path):
    received: list[dict[str, object]] = []

    class RecordingConverter:
        def convert(self, source: Path, **kwargs: object):
            received.append(kwargs)
            return SimpleNamespace(markdown=f"# {source.stem}")

    xlsx = tmp_path / "颜色.xlsx"
    xlsx.write_bytes(b"test")
    docx = tmp_path / "普通.docx"
    docx.write_bytes(b"test")

    convert_file(xlsx, preserve_excel_colors=True, converter_factory=RecordingConverter)
    convert_file(docx, preserve_excel_colors=True, converter_factory=RecordingConverter)

    assert received == [{"preserve_excel_colors": True}, {}]


def test_existing_output_is_preserved_when_overwrite_is_disabled(tmp_path: Path):
    source = tmp_path / "example.pdf"
    source.write_bytes(b"test")
    destination = tmp_path / "example.md"
    destination.write_text("keep", encoding="utf-8")

    with pytest.raises(OutputExistsError):
        convert_file(source, overwrite=False, converter_factory=FakeConverter)

    assert destination.read_text(encoding="utf-8") == "keep"


def test_unsupported_file_is_rejected(tmp_path: Path):
    source = tmp_path / "archive.rar"
    source.write_bytes(b"test")

    with pytest.raises(UnsupportedFileError):
        convert_file(source, converter_factory=FakeConverter)


def test_xlsx_preserves_merged_hierarchy_and_text_identifiers(tmp_path: Path):
    source = tmp_path / "层级清单.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "功能清单"
    sheet.append(["功能模块", "编号", "功能项", "说明\n（原文）"])
    sheet.append(["数据建模", "1.1", "多模型管理", "第一项\n补充说明"])
    sheet.append([None, "1.2", "模型生命周期", "第二项"])
    sheet.append([None, "1.10", "敏感字段标识", "第十项"])
    sheet.merge_cells("A2:A4")
    workbook.save(source)

    destination = convert_file(source)
    markdown = destination.read_text(encoding="utf-8")

    assert "| 功能模块 | 编号 | 功能项 | 说明 （原文） |" in markdown
    assert "| 数据建模 | 1.1 | 多模型管理 | 第一项； 补充说明 |" in markdown
    assert "| 数据建模 | 1.2 | 模型生命周期 | 第二项 |" in markdown
    assert "| 数据建模 | 1.10 | 敏感字段标识 | 第十项 |" in markdown
    assert "1.20" not in markdown
    assert "NaN" not in markdown
    assert "Unnamed:" not in markdown
    assert "<br>" not in markdown


def test_xlsx_cell_colors_are_optional_and_default_to_unchanged_output(tmp_path: Path):
    source = tmp_path / "颜色可选.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["评价", "结果"])
    sheet.append(["雅思", "6.5"])
    sheet["A2"].fill = PatternFill(fill_type="solid", fgColor="FFFF0000")
    sheet["B2"].fill = PatternFill(fill_type="solid", fgColor="FF92D050")
    workbook.save(source)

    markdown = convert_file(source).read_text(encoding="utf-8")

    assert "| 雅思 | 6.5 |" in markdown
    assert "Excel 单元格颜色图例" not in markdown
    assert "🟥" not in markdown
    assert "🟩" not in markdown
    assert "字体、颜色或版式" in markdown


def test_xlsx_cell_colors_emit_emoji_and_exact_hex_legend_when_enabled(tmp_path: Path):
    source = tmp_path / "颜色保留.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["红色", "橙色", "绿色", "主题色", "索引色"])
    sheet.append(["0-4.5", "6", "6.5", "主题", "索引"])
    sheet["A2"].fill = PatternFill(fill_type="solid", fgColor="FFFF0000")
    sheet["B2"].fill = PatternFill(fill_type="solid", fgColor="FFED7D31")
    sheet["C2"].fill = PatternFill(fill_type="solid", fgColor="FF92D050")
    sheet["D2"].fill = PatternFill(fill_type="solid", fgColor=Color(theme=5))
    sheet["E2"].fill = PatternFill(fill_type="solid", fgColor=Color(indexed=5))
    workbook.save(source)

    markdown = convert_file(source, preserve_excel_colors=True).read_text(encoding="utf-8")

    assert "| 🟥 0-4.5 | 🟧 6 | 🟩 6.5 |" in markdown
    assert "## Excel 单元格颜色图例" in markdown
    assert "`#FF0000`" in markdown
    assert "`#ED7D31`" in markdown
    assert "`#92D050`" in markdown
    assert "`#C0504D`" in markdown
    assert "`#FFFF00`" in markdown
    assert "单元格填充色：5/5 个非空着色单元格已标记（5 种精确颜色）" in markdown
    assert "不推断颜色的业务含义" in markdown


def test_xlsx_custom_colors_with_same_emoji_receive_stable_collision_ids(tmp_path: Path):
    source = tmp_path / "自定义颜色.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["方案 A", "方案 B"])
    sheet.append(["第一项", "第二项"])
    sheet["A2"].fill = PatternFill(fill_type="solid", fgColor="FFF00000")
    sheet["B2"].fill = PatternFill(fill_type="solid", fgColor="FFE01020")
    workbook.save(source)

    markdown = convert_file(source, preserve_excel_colors=True).read_text(encoding="utf-8")

    assert "| 🟥[C1] 第一项 | 🟥[C2] 第二项 |" in markdown
    assert "- 🟥[C1] = `#F00000`" in markdown
    assert "- 🟥[C2] = `#E01020`" in markdown


def test_xlsx_preserves_nested_lists_as_coordinate_linked_markdown_details(tmp_path: Path):
    source = tmp_path / "结构化列表.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "功能清单"
    sheet.append(["编号", "一级模块", "二级模块", "功能项", "说明"])
    sheet.append(
        [
            25,
            "招聘过程",
            "简历筛选",
            "智能筛选",
            "1. AI智能匹配\n• 匹配度评分\n• 高匹配度标识\n2. AI自动推荐\n• 群星推荐\n• 自动筛选",
        ]
    )
    workbook.save(source)

    markdown = convert_file(source).read_text(encoding="utf-8")

    assert "| 25 | 招聘过程 | 简历筛选 | 智能筛选 | 详见「E2 结构化详情」 |" in markdown
    assert "#### E2 · 25 / 招聘过程 / 简历筛选 / 智能筛选" in markdown
    assert "1. AI智能匹配\n   - 匹配度评分\n   - 高匹配度标识" in markdown
    assert "2. AI自动推荐\n   - 群星推荐\n   - 自动筛选" in markdown
    assert "结构化列表单元格：1 个已保留为 Markdown 层级" in markdown
    assert "<br>" not in markdown
    assert "<ol" not in markdown
    assert "<ul" not in markdown


def test_xlsx_renders_horizontal_merges_as_sections_and_notes(tmp_path: Path):
    source = tmp_path / "报价表.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "成本总表"
    sheet.append(["序号", "需求类型", "金额", "价值"])
    sheet.append(["基础系统包"])
    sheet.merge_cells("A2:D2")
    sheet.append([1, "标品", 100, "基础能力"])
    sheet.append(["说明"])
    sheet.merge_cells("A4:D4")
    sheet.append(["报价不含第三方软件许可。"])
    sheet.merge_cells("A5:D6")
    workbook.save(source)

    markdown = convert_file(source).read_text(encoding="utf-8")

    assert "### 基础系统包" in markdown
    assert "| 1 | 标品 | 100 | 基础能力 |" in markdown
    assert "### 说明" in markdown
    assert "> 报价不含第三方软件许可。" in markdown


def test_xlsx_folds_merge_only_rows_without_filling_unmerged_identifiers(tmp_path: Path):
    source = tmp_path / "合并延续行.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "功能清单"
    sheet.append(["编号", "一级", "二级", "功能", "说明"])
    sheet.append([67, "AI组件", "效率工具", "全网简历搜索", "搜索所有渠道"])
    sheet.append([None, None, None, None, None])
    for column in "ABCDE":
        sheet.merge_cells(f"{column}2:{column}3")
    sheet.append([68, "AI组件", "效率工具", "全天候人才跟盯", "自动通知"])
    sheet.append([52, "招聘结果", "报表", "BI系统集成", "第一项"])
    sheet.append([None, None, None, "内建报表", "第二项"])
    for column in "ABC":
        sheet.merge_cells(f"{column}5:{column}6")
    sheet.append([None, "其他", "配置", "未编号", "第三项"])
    workbook.save(source)

    markdown = convert_file(source).read_text(encoding="utf-8")

    assert markdown.count("| 67 | AI组件 | 效率工具 | 全网简历搜索 | 搜索所有渠道 |") == 1
    assert "| 68 | AI组件 | 效率工具 | 全天候人才跟盯 | 自动通知 |" in markdown
    assert "| 52 | 招聘结果 | 报表 | BI系统集成 | 第一项 |" in markdown
    assert "| 52 | 招聘结果 | 报表 | 内建报表 | 第二项 |" in markdown
    assert "|  | 其他 | 配置 | 未编号 | 第三项 |" in markdown
    assert "已折叠 1 个仅用于合并单元格排版的延续行（源表第 3 行）" in markdown


def test_xlsx_warns_only_for_duplicate_source_identifiers_within_a_section(tmp_path: Path):
    source = tmp_path / "重复编号.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "功能清单"
    sheet.append(["编号", "功能项"])
    sheet.append([1, "甲"])
    sheet.append([2, "乙"])
    sheet.append([2, "丙"])
    workbook.save(source)

    markdown = convert_file(source).read_text(encoding="utf-8")

    assert "源数据提示" in markdown
    assert "`2`（A3、A4）" in markdown
    assert "DropMD 已按原文件保留" in markdown


def test_xlsx_ignores_styled_empty_tail(tmp_path: Path):
    source = tmp_path / "有效区域.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["名称", "金额"])
    sheet.append(["项目", 100])
    sheet["H41"].fill = copy(sheet["A1"].fill)
    workbook.save(source)

    markdown = convert_file(source).read_text(encoding="utf-8")

    assert "| 名称 | 金额 |" in markdown
    assert "列 H" not in markdown
    assert markdown.count("| 项目 | 100 |") == 1


def test_xlsx_emits_workbook_semantics_and_formula_provenance(tmp_path: Path):
    source = tmp_path / "语义报价.xlsx"
    workbook = Workbook()
    summary = workbook.active
    summary.title = "报价汇总"
    summary.append(["项目", "金额"])
    summary.append(["合计", "=SUM(B3:B4)"])
    summary.append(["一期", 10])
    summary.append(["二期", 20])
    details = workbook.create_sheet("隐藏明细")
    details.append(["编号", "说明"])
    details.append(["1.10", "保留显示编号"])
    details.sheet_state = "hidden"
    workbook.save(source)
    set_formula_cache(source, "B2", "30")

    markdown = convert_file(source).read_text(encoding="utf-8")

    assert markdown.startswith("# 语义报价\n")
    assert "2 个工作表（1 个可见，1 个隐藏）" in markdown
    assert "1 个公式单元格" in markdown
    assert "## 语义转换完整性" in markdown
    assert "工作表：2/2 已保留（隐藏工作表 1/1）" in markdown
    assert "可追踪语义单元格：12/12 已表示" in markdown
    assert "公式：1/1 已保留（缓存结果 1/1）" in markdown
    assert "合成记录：0" in markdown
    assert "编号显示值变更：0" in markdown
    assert "- `报价汇总`：区域 `A1:B4`；可见；1 个公式" in markdown
    assert "> **源表提示：** 原工作表为隐藏状态。" in markdown
    assert "| **合计** | **30** |" in markdown
    assert "- `B2`：`=SUM(B3:B4)` → `30`" in markdown
